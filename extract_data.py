import openpyxl, json

wb = openpyxl.load_workbook(r'c:\Users\K7813444\OneDrive - Saint-Gobain\Desktop\2026\ACE\SG AIS GGL_TOOL.xlsx')

all_data = []

def norm_glazing(g):
    if g is None: return ''
    g = g.strip()
    if g.lower() == 'single glazing': return 'SGU'
    if g.lower() == 'double glazing': return 'DGU'
    return g

def norm_shade(s):
    if s is None: return ''
    return s.strip()

def safe(v):
    if v is None: return 0
    try: return float(v)
    except: return 0

sheets = [
    ('SG_EN', 'Saint-Gobain'),
    ('SG_NFRC', 'Saint-Gobain'),
    ('GGL_NFRC', 'Guardian'),
    ('AIS_EN', 'Asahi'),
]

for sheet_name, brand in sheets:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[3] is None: continue
        all_data.append({
            'Brand': brand,
            'Standard': str(row[0]).strip() if row[0] else '',
            'Shade': norm_shade(row[1]),
            'GlazingType': norm_glazing(row[2]),
            'ProductName': str(row[3]).strip(),
            'VLT': safe(row[4]),
            'ER': safe(row[5]),
            'IR': safe(row[6]),
            'SHGC': safe(row[7]),
            'UValue': safe(row[8]),
        })

with open(r'c:\Users\K7813444\OneDrive - Saint-Gobain\Desktop\2026\ACE\products.json', 'w') as f:
    json.dump(all_data, f)

print(f"Exported {len(all_data)} products")
for b in ['Saint-Gobain', 'Guardian', 'Asahi']:
    count = len([d for d in all_data if d['Brand'] == b])
    print(f"  {b}: {count}")

shades = sorted(set(d['Shade'] for d in all_data))
print(f"Shades: {shades}")
glazing = sorted(set(d['GlazingType'] for d in all_data))
print(f"Glazing: {glazing}")
standards = sorted(set(d['Standard'] for d in all_data))
print(f"Standards: {standards}")
